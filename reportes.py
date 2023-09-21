from datetime import date, timedelta
import os
import sqlite3
import sys
import tkinter as tk
from tkinter import *
from tkinter import ttk
import openpyxl

class reportes(Tk):
    def __init__(self):
        super().__init__()
        self.title("Reportes")
        self.geometry("800x600")
        self.resizable(True, True)
        self.decoracionesReportes()



    def decoracionesReportes(self):
        ########################### CONTENEDORES ###########################
        # Encabezado de página
        frame1 = Frame(self, bg='#B5B5B5')
        frame1.pack(side=TOP, fill=X)


        #Frame texto
        frame2 = Frame(self, bg='#B5B5B5', width=500)#D9D9D9
        frame2.pack(side=LEFT, fill=BOTH, padx=10, pady=10)
        # RESUMEN
        resumen = Frame(frame2, bg='#B5B5B5')
        resumen.pack(anchor='n', fill= Y,padx=5, pady=10)#side=TOP
        subR1 =Frame(resumen, bg='#B5B5B5')#B5B5B5
        subR1.pack(side=BOTTOM, padx=5, pady=5, fill=X)
        subR2 = Frame(resumen, bg='#B5B5B5')
        subR2.pack(side=BOTTOM, padx=5, pady=5, fill=X)
        subR3 = Frame(resumen, bg='#B5B5B5')
        subR3.pack(side=BOTTOM, padx=5, pady=5, fill=X)
        excel = Frame(frame2, bg='#B5B5B5')
        excel.pack(side=BOTTOM, padx=5, pady=5, fill=X)


        # Ventas
        frame3 = Frame(frame2, bg='#B5B5B5')
        frame3.pack(anchor='n', fill=Y,padx=10, pady=10)
        # frame3.pack(side=TOP, padx=16, pady=10)
        #Subcontenedores
        subV1 = Frame(frame3, bg='#B5B5B5')#B5B5B5
        subV1.pack(side=BOTTOM, padx=5, pady=5)
        subV2 = Frame(frame3, bg='#B5B5B5')
        subV2.pack(side=BOTTOM, padx=5, pady=5)


        # Articulos por agotarse
        frame4 = Frame(self, bg='#B5B5B5')
        frame4.pack(side=RIGHT, padx=10, pady=10, fill=Y)


        ######################## BOTONES ########################
        # Botón Regresar
        self.bRegresar = Button(frame1, text='Regresar', command=self.fRegresar, bg='#D9D9D9', fg='Black', 
                                        font=("Arial Black",9), highlightbackground='#D9D9D9'
                                        , width=10, height=2)
        self.bRegresar.pack(side=RIGHT, padx=10, pady=5)

        bActualizar = Button(excel, text='Actualizar', command=self.actualizar, bg='#088423',fg='white', 
                                        font=("Arial Black",9), highlightbackground='#D9D9D9'
                                        , width=20, height=2)
        bActualizar.pack()

        bExportar = Button(excel, text='Exportar a Excel', command=self.exportar_a_excel, background='#088423', fg='white',
                            font=("Arial Black", 9), highlightbackground='#D9D9D9'
                                        , width=20, height=2)
        bExportar.pack()
        ######################## Entradas y salidas ########################
        # Texto Reportes
        self.lb1 = Label(frame1, text="Reportes", bg='#B5B5B5', font=("Arial Black", 50))
        self.lb1.pack(side=TOP, padx=10, pady=10)


        #Resumen
        self.titulo = Label(resumen, text='Resumen', background='#B5B5B5', font=("Arial Black", 10))
        self.titulo.pack(side=TOP)
        # Venta del día
        self.dia = Label(subR1, text='Venta del día: ', bg='#B5B5B5', font=("Arial Black", 10))
        self.dia.pack(side=LEFT)#anchor=W   padx=10, pady=5
        self.diaE = Entry(subR1)
        self.diaE.pack(side=LEFT, fill=X, expand=True) #, padx=10, pady=10
        #Venta semanal
        self.semana = Label(subR2, text='\tTotal venta semanal: ', bg='#B5B5B5', font=("Arial Black", 10))
        self.semana.pack(side=LEFT)#anchor=W,  padx=10, pady=5
        self.semanaE = Entry(subR2)
        self.semanaE.pack(side=LEFT, fill=X, expand=True) #, padx=5
        # Venta mensual
        self.mensual = Label(subR3, text='\tTotal venta mensual: ', bg='#B5B5B5', font=("Arial Black", 10))
        self.mensual.pack(side=LEFT)#anchor=W, padx=10, pady=5
        self.mensualE = Entry(subR3)
        self.mensualE.pack(side=LEFT, fill=X, expand=True)#, padx=5



        # Más vendido de la semana
        self.masVendido = Label(subV1, text='Más vendido de la semana:', bg='#B5B5B5', font=("Arial Black", 10))
        self.masVendido.pack(side=LEFT, anchor=W, padx=10, pady=5)
        self.masVendidoE = Entry(subV1)
        self.masVendidoE.pack(side=LEFT, padx=5)
        # Menos vendido de la semana
        self.mVendido = Label(subV2, text='Menos vendido de la semana:', bg='#B5B5B5', font=("Arial Black", 10))
        self.mVendido.pack(side=LEFT, anchor=W, padx=10, pady=5)
        self.mVendidoE = Entry(subV2)
        self.mVendidoE.pack(side=LEFT, padx=5)


        #Artículos por agotarse
        self.aAgotarse = Label(frame4, text='Artículos por agotarse', bg='#B5B5B5', font=('Arial Black', 10), anchor=CENTER)
        self.aAgotarse.pack(side=TOP, padx=10, pady=5)


        ######################## TABLAS ########################
        scrollbar = ttk.Scrollbar(frame4)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tabla = ttk.Treeview(frame4, yscrollcommand=scrollbar.set)
        self.tabla.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.tabla.yview)
        
        self.tabla['columns'] = ('Producto', 'Cantidad', 'Costo')
        self.tabla.column('#0', width=0, stretch=tk.NO)
        self.tabla.column('Producto', anchor=tk.CENTER, width=100)
        self.tabla.column('Cantidad', anchor=tk.CENTER, width=100)
        self.tabla.column('Costo', anchor=tk.CENTER, width=100)

        self.tabla.heading('#0', text='', anchor=tk.CENTER)
        self.tabla.heading('Producto', text='Producto', anchor=tk.CENTER)
        self.tabla.heading('Cantidad', text='Cantidad', anchor=tk.CENTER)
        self.tabla.heading('Costo', text='Costo', anchor=tk.CENTER)


    def fRegresar(self):
        self.withdraw()
        os.system(f"{sys.executable} main.py")
    
    def actualizar(self):
        self.mostrar_productos_proximos_agotarse()
        self.mostrar_productos_mas_vendidos()
        
        self.mostrar_articulo_mas_vendido_semana()

    def mostrar_productos_mas_vendidos(self):
        # Conectar a la base de datos
        conn = sqlite3.connect('regalitosdb.db')
        c = conn.cursor()

        # Obtener la fecha actual
        fecha_actual = date.today().strftime("%Y-%m-%d")
        # Obtener el total de la venta del día
        c.execute("SELECT SUM(total_v) FROM concepto_venta WHERE fecha = ?", (fecha_actual,))
        resultado = c.fetchone()
        total_venta_dia = resultado[0] if resultado[0] else 0

        self.diaE.delete(0, tk.END)
        self.diaE.insert(0, f"({fecha_actual}): {total_venta_dia}")
        self.diaE.config(width=len(f"({fecha_actual}), {total_venta_dia}"))


        # Obtener la fecha actual y la fecha hace una semana
        fecha_actual = date.today()
        fecha_semana_pasada = fecha_actual - timedelta(days=7)

        # Obtener el total de la venta de la semana
        c.execute("SELECT SUM(total_v) FROM concepto_venta WHERE fecha >= ? AND fecha <= ?", (fecha_semana_pasada, fecha_actual))
        resultado = c.fetchone()
        total_venta_semana = resultado[0] if resultado[0] else 0

        self.semanaE.delete(0, tk.END)
        self.semanaE.insert(0, f"({fecha_actual}): {total_venta_semana}")
        self.semanaE.config(width=len(f"                 : {total_venta_semana}"))#f"({fecha_actual}), {total_venta_}")


        # Obtener el mes actual
        mes_actual = date.today().strftime("%m")

        # Imprimir la consulta SQL y los parámetros para verificarlos
        print("Consulta SQL:", "SELECT SUM(total_v) FROM concepto_venta WHERE strftime('%m', fecha) = ?", (mes_actual,))

        # Obtener el total de la venta del mes
        c.execute("SELECT SUM(total_v) FROM concepto_venta WHERE strftime('%Y-%m', fecha) = ?", (fecha_actual.strftime("%Y-%m"),))
        # c.execute("SELECT SUM(total_v) FROM concepto_venta WHERE strftime('%m', fecha) = ?", (mes_actual,))
        resultado = c.fetchone()
        total_venta_mes = resultado[0] if resultado[0] else 0

        # Imprimir el resultado de la consulta
        print("Resultado:", resultado)

        self.mensualE.delete(0, tk.END)
        self.mensualE.insert(0, f"      {total_venta_mes}")
        self.mensualE.config(width=len(f"       {total_venta_mes}"))

        conn.close()


        # # Obtener el mes actual
        # mes_actual = date.today().strftime("%m")

        # # Obtener el total de la venta del mes
        # c.execute("SELECT SUM(total_v) FROM concepto_venta WHERE strftime('%m', fecha) = ?", (mes_actual,))
        # resultado = c.fetchone()
        # total_venta_mes = resultado[0] if resultado[0] else 0
        # self.mensualE.delete(0, tk.END)
        # self.mensualE.insert(0, f"      {total_venta_mes}")
        # self.mensualE.config(width=len(f"       {total_venta_mes}"))

        # conn.close()

    
    def mostrar_productos_proximos_agotarse(self):
        # Conectar a la base de datos
        conn = sqlite3.connect('regalitosdb.db')
        c = conn.cursor()

        # Obtener los productos próximos a agotarse
        c.execute("SELECT nombre, existencias, costo FROM productos WHERE existencias <= 4")
        resultados = c.fetchall()

        # Limpiar la tabla antes de agregar los nuevos datos
        self.tabla.delete(*self.tabla.get_children())

        # Agregar los productos a la tabla
        for resultado in resultados:
            self.tabla.insert('', 'end', values=resultado)

        # Cerrar la conexión con la base de datos
        conn.close()


    def mostrar_articulo_mas_vendido_semana(self):
        # Conectar a la base de datos
        conn = sqlite3.connect('regalitosdb.db')
        c = conn.cursor()

        # Obtener el código de barras del artículo más vendido de la semana
        c.execute("SELECT art_vendido, SUM(cant_art) FROM concepto_venta WHERE strftime('%W', fecha) = strftime('%W', 'now') GROUP BY art_vendido ORDER BY SUM(cant_art) DESC LIMIT 1")
        resultado = c.fetchone()
        codigo_barras, cantidad_semana = resultado if resultado else ('', 0)

        # Obtener el nombre del producto correspondiente al código de barras
        c.execute("SELECT nombre FROM productos WHERE codigo = ?", (codigo_barras,))
        resultado_producto = c.fetchone()
        producto_semana = resultado_producto[0] if resultado_producto else ''

        # Actualizar el Entry con el nombre del producto y la cantidad vendida
        texto = f"{producto_semana} ({cantidad_semana} unidades)"
        self.masVendidoE.delete(0, tk.END)
        self.masVendidoE.insert(0, texto)
        self.masVendidoE.config(width=len(texto))

        # Cerrar la conexión con la base de datos
        conn.close()

    def exportar_a_excel(self):
        # Obtener los datos de los Entry
        datos_dia = self.diaE.get()
        datos_semana = self.semanaE.get()
        datos_mes = self.mensualE.get()

        # Crear una instancia de ExportadorExcel
        exportador = ExportadorExcel("informe.xlsx")

        # Exportar los datos a Excel
        exportador.exportar([datos_dia], [datos_semana], [datos_mes])



class ExportadorExcel:
    def __init__(self, archivo):
        self.archivo = archivo

    def exportar(self, datos_dia, datos_semana, datos_mes):
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Añadir títulos a las columnas
        hoja.cell(row=1, column=1, value='Venta del día')
        hoja.cell(row=1, column=2, value='Venta semanal')
        hoja.cell(row=1, column=3, value='Venta mensual')

        # Escribir los datos en la hoja de cálculo
        for fila, dato in enumerate(datos_dia, start=2):
            hoja.cell(row=fila, column=1, value=dato)

        for fila, dato in enumerate(datos_semana, start=2):
            hoja.cell(row=fila, column=2, value=dato)

        for fila, dato in enumerate(datos_mes, start=2):
            hoja.cell(row=fila, column=3, value=dato)

        # Guardar el libro en el archivo especificado
        libro.save(self.archivo)




#     def exportar_a_excel(self):
#         # Obtener los datos de los Entry
#         datos_dia = self.diaE.get()
#         datos_semana = self.semanaE.get()
#         datos_mes = self.mensualE.get()

#         # Crear una instancia de ExportadorExcel
#         exportador = ExportadorExcel("informe.xlsx")

#         # Exportar los datos a Excel
#         exportador.exportar([datos_dia, datos_semana, datos_mes])


# class ExportadorExcel:
#     def __init__(self, archivo):
#         self.archivo = archivo

#     def exportar(self, datos):
#         # Crear un nuevo libro de Excel
#         libro = openpyxl.Workbook()
#         hoja = libro.active

#         # Escribir los datos en la hoja de cálculo
#         for fila, dato in enumerate(datos, start=1):
#             hoja.cell(row=fila, column=1, value=dato)

#         # Guardar el libro en el archivo especificado
#         libro.save(self.archivo)


if __name__ == "__main__":
    root = reportes()
    root.mainloop()
